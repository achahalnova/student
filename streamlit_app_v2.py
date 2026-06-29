import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import io
import os

st.set_page_config(page_title="Student Excel Analyzer", page_icon="📊", layout="wide")

# Sidebar Navigation instead of Tabs (much more stable for older Streamlit versions)
st.sidebar.title("Navigation")
page = st.sidebar.radio("Go to", ["Results Analyzer (Excel)", "Exam Grader (.dat)"])

if page == "Results Analyzer (Excel)":
    st.title("Student Results Analyzer 📊")
    # Removed the word 'Aakash' as requested
    st.write("Upload your weekly Excel file. This app will instantly filter the NSPIRA-CC branch, format the ranks (Top 33% = Green, Middle = Yellow, Bottom = Red), and give you the clean file!")
    
    exam_type = st.radio("Select Exam Type:", ["Engineering (JEE)", "Medical (NEET)"])
    uploaded_file = st.file_uploader("Upload Student Results Excel", type=["xlsx", "xls", "csv"])
    
    if uploaded_file is not None:
        st.info("Reading file...")
        
        try:
            if uploaded_file.name.lower().endswith(".csv"):
                df_raw = pd.read_csv(uploaded_file, header=None)
            else:
                if exam_type == "Medical (NEET)":
                    try:
                        df_raw = pd.read_excel(uploaded_file, sheet_name='EST', header=None)
                    except Exception:
                        st.warning("Could not find 'EST' sheet. Falling back to the first sheet...")
                        df_raw = pd.read_excel(uploaded_file, header=None)
                else:
                    df_raw = pd.read_excel(uploaded_file, header=None)
            
            if len(df_raw) < 3:
                st.error("Error: Excel file does not contain enough rows to analyze.")
                st.stop()
                
            header_row_idx = 1
            for idx in range(min(15, len(df_raw))):
                row_vals = [str(x).upper().strip() for x in df_raw.iloc[idx].tolist()]
                if 'NAME' in row_vals or 'STUDENT NAME' in row_vals or 'BRANCH' in row_vals:
                    header_row_idx = idx
                    break
                    
            data_rows = df_raw.iloc[header_row_idx + 1:]
            header_row = [str(x).strip() for x in df_raw.iloc[header_row_idx].tolist()]
    
            def get_first_idx(col_name_or_list, start_idx=0, required=True):
                col_names = [str(c).upper().strip() for c in (col_name_or_list if isinstance(col_name_or_list, list) else [col_name_or_list])]
                for i in range(start_idx, len(header_row)):
                    if str(header_row[i]).upper().strip() in col_names:
                        return i
                if required:
                    available_cols = ", ".join([str(x).strip() for x in header_row if str(x).strip() and str(x).strip().lower() != 'nan'])
                    st.error(f"Error: Could not find column '{col_name_or_list}' in the Excel header. Available columns in detected header row: {available_cols}")
                    st.stop()
                return None
                
            branch_idx = get_first_idx(['BRANCH', 'CAMPUS', 'CENTRE'], required=False)
            name_idx = get_first_idx(['NAME', 'STUDENT NAME'])
            
            pm_idx = get_first_idx('PM', 0)
            pr_idx = get_first_idx('PR', pm_idx)
            pw_idx = get_first_idx('W', pm_idx)
            
            cm_idx = get_first_idx('CM', 0)
            cr_idx = get_first_idx('CR', cm_idx)
            cw_idx = get_first_idx('W', cm_idx)
            
            if exam_type == "Medical (NEET)":
                bm_idx = get_first_idx('BM', 0)
                br_idx = get_first_idx('BR', bm_idx)
                bw_idx = get_first_idx('W', bm_idx)
                
                zm_idx = get_first_idx('ZM', 0)
                zr_idx = get_first_idx('ZR', zm_idx)
                zw_idx = get_first_idx('W', zm_idx)
            else:
                mm_idx = get_first_idx('MM', 0)
                mr_idx = get_first_idx('MR', mm_idx)
                mw_idx = get_first_idx('W', mm_idx)
                
            tm_idx = get_first_idx(['TM', 'Total'], 0)
            tr_idx = get_first_idx('TR', tm_idx)
    
            if branch_idx is not None:
                all_branches = data_rows[branch_idx].dropna().unique().tolist()
                mh_mum_branches = [str(b) for b in all_branches if str(b).startswith("MH-MUM")]
                
                branch_options = mh_mum_branches if mh_mum_branches else [str(b) for b in all_branches]
                
                if not branch_options:
                    st.warning("No branches were found in the uploaded file. Showing all students.")
                    df_filtered = data_rows.copy()
                else:
                    target_branch = st.selectbox("Select branch to filter by:", options=branch_options)
                    st.info(f"Filtering for branch {target_branch}...")
                    df_filtered = data_rows[data_rows[branch_idx] == target_branch].copy()
                    if df_filtered.empty:
                        st.warning(f"No students found for branch '{target_branch}'.")
                        st.stop()
            else:
                st.info("No BRANCH column found. Showing all students.")
                df_filtered = data_rows.copy()
    
            if exam_type == "Medical (NEET)":
                req_indices = [name_idx, pm_idx, pr_idx, pw_idx, cm_idx, cr_idx, cw_idx, bm_idx, br_idx, bw_idx, zm_idx, zr_idx, zw_idx, tm_idx, tr_idx]
                df_final = df_filtered.iloc[:, req_indices].copy()
                df_final.columns = ['NAME', 'PM', 'PR', 'Physics W', 'CM', 'CR', 'Chemistry W', 'BM', 'BR', 'Botany W', 'ZM', 'ZR', 'Zoology W', 'TM', 'TR']
                cols_to_convert = ['PM', 'PR', 'Physics W', 'CM', 'CR', 'Chemistry W', 'BM', 'BR', 'Botany W', 'ZM', 'ZR', 'Zoology W', 'TM', 'TR']
            else:
                req_indices = [name_idx, mm_idx, mr_idx, mw_idx, pm_idx, pr_idx, pw_idx, cm_idx, cr_idx, cw_idx, tm_idx, tr_idx]
                df_final = df_filtered.iloc[:, req_indices].copy()
                df_final.columns = ['NAME', 'MM', 'MR', 'Math W', 'PM', 'PR', 'Physics W', 'CM', 'CR', 'Chemistry W', 'TM', 'TR']
                cols_to_convert = ['MM', 'MR', 'Math W', 'PM', 'PR', 'Physics W', 'CM', 'CR', 'Chemistry W', 'TM', 'TR']
    
            df_final = df_final.reset_index(drop=True)
    
            for col in cols_to_convert:
                df_final[col] = pd.to_numeric(df_final[col], errors='coerce')
    
            def get_tertiles_marks(local_series, global_series):
                s_local = local_series.dropna()
                s_global = global_series.dropna()
                
                if len(s_local) == 0 or len(s_global) == 0: 
                    return [], [], []
                    
                try:
                    p33 = s_global.quantile(0.33)
                    p66 = s_global.quantile(0.66)
                    
                    worst_indices = s_local[s_local < p33].index.tolist()
                    avg_indices = s_local[(s_local >= p33) & (s_local < p66)].index.tolist()
                    best_indices = s_local[s_local >= p66].index.tolist()
                except:
                    return [], [], []
                    
                return best_indices, worst_indices, avg_indices
    
            if branch_idx is not None:
                is_co_branch = data_rows[branch_idx].astype(str).str.upper().str.contains('CO')
                data_rows_global = data_rows[~is_co_branch].copy()
            else:
                data_rows_global = data_rows.copy()
                
            if exam_type == "Medical (NEET)":
                df_global = data_rows_global.iloc[:, req_indices].copy()
                df_global.columns = ['NAME', 'PM', 'PR', 'Physics W', 'CM', 'CR', 'Chemistry W', 'BM', 'BR', 'Botany W', 'ZM', 'ZR', 'Zoology W', 'TM', 'TR']
            else:
                df_global = data_rows_global.iloc[:, req_indices].copy()
                df_global.columns = ['NAME', 'MM', 'MR', 'Math W', 'PM', 'PR', 'Physics W', 'CM', 'CR', 'Chemistry W', 'TM', 'TR']
    
            for col in cols_to_convert:
                df_global[col] = pd.to_numeric(df_global[col], errors='coerce')
                
            df_top500 = df_global.sort_values(by='TM', ascending=False).head(500)
    
            st.info("Calculating percentiles based on top 500 non-CO students' marks and coloring...")
            best_p, worst_p, avg_p = get_tertiles_marks(df_final['PM'], df_top500['PM'])
            best_c, worst_c, avg_c = get_tertiles_marks(df_final['CM'], df_top500['CM'])
            best_t, worst_t, avg_t = get_tertiles_marks(df_final['TM'], df_top500['TM'])
    
            if exam_type == "Medical (NEET)":
                best_b, worst_b, avg_b = get_tertiles_marks(df_final['BM'], df_top500['BM'])
                best_z, worst_z, avg_z = get_tertiles_marks(df_final['ZM'], df_top500['ZM'])
            else:
                best_m, worst_m, avg_m = get_tertiles_marks(df_final['MM'], df_top500['MM'])
    
            output_buffer = io.BytesIO()
            df_final.to_excel(output_buffer, index=False)
            output_buffer.seek(0)
            
            wb = openpyxl.load_workbook(output_buffer)
            ws = wb.active
    
            ws.column_dimensions['A'].width = 25
    
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
            def apply_color(ws, col_name, best_indices, worst_indices, avg_indices):
                col_idx = None
                for cell in ws[1]:
                    if cell.value == col_name:
                        col_idx = cell.column
                        break
                if not col_idx: return
                
                for idx in best_indices: ws.cell(row=idx+2, column=col_idx).fill = green_fill
                for idx in avg_indices: ws.cell(row=idx+2, column=col_idx).fill = yellow_fill
                for idx in worst_indices: ws.cell(row=idx+2, column=col_idx).fill = red_fill
    
            apply_color(ws, 'PM', best_p, worst_p, avg_p)
            apply_color(ws, 'CM', best_c, worst_c, avg_c)
            apply_color(ws, 'TM', best_t, worst_t, avg_t)
    
            if exam_type == "Medical (NEET)":
                apply_color(ws, 'BM', best_b, worst_b, avg_b)
                apply_color(ws, 'ZM', best_z, worst_z, avg_z)
            else:
                apply_color(ws, 'MM', best_m, worst_m, avg_m)
    
            final_buffer = io.BytesIO()
            wb.save(final_buffer)
            final_buffer.seek(0)
            
            st.success("Analysis Complete! Your file is ready to download.")
            
            st.write("### Data Preview (First 5 Rows)")
            st.dataframe(df_final.head())
            
            base_name, _ = os.path.splitext(uploaded_file.name)
            new_filename = f"{base_name}_Analyzed.xlsx"
            st.download_button(
                label="⬇️ Download Formatted Excel File",
                data=final_buffer,
                file_name=new_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
        except Exception as e:
            st.error(f"An error occurred while processing the file: {e}")

elif page == "Exam Grader (.dat)":
    st.title("Exam Grader (.dat) 📝")
    st.write("Upload student .dat files and an Answer Key to process scores.")
    
    batch_options = [
        "junior jee 1", "junior jee 2", "junior jee 3", "junior jee 4", "junior jee 5",
        "junior neet 1", "junior neet 2", "junior neet 3",
        "senior jee 1", "senior jee 2", "senior jee 3", "senior jee 4", "senior jee 5",
        "senior neet 1", "senior neet 2",
        "start neet"
    ]
    
    selected_batch = st.selectbox("Select Batch:", options=batch_options)
    
    st.write("### Upload Files")
    col1, col2 = st.columns(2)
    with col1:
        dat_file = st.file_uploader("Upload Student Data File (.dat/.iit or extensionless)", type=None)
    with col2:
        answer_key_file = st.file_uploader("Upload Answer Key", type=["docx", "doc", "xlsx", "xls"])
    
    if dat_file is not None and answer_key_file is not None:
        st.info("Files uploaded successfully.")
        
        def extract_answer_key_from_docx(docx_file):
            import docx
            import pandas as pd
            
            doc = docx.Document(docx_file)
            data = []
        
            for table_idx, table in enumerate(doc.tables):
                if table_idx == 0:
                    first_cell_text = table.rows[0].cells[0].text.strip()
                    if not first_cell_text.isdigit():
                        continue
        
                for i in range(0, len(table.rows), 2):
                    if i + 1 >= len(table.rows):
                        break
                    q_row = table.rows[i]
                    a_row = table.rows[i+1]
                    
                    for q_cell, a_cell in zip(q_row.cells, a_row.cells):
                        q_text = q_cell.text.strip()
                        a_text = a_cell.text.strip()
                        
                        if q_text.isdigit():
                            data.append({
                                "Question Number": int(q_text),
                                "Answer": a_text
                            })
        
            if not data:
                return None
        
            data.sort(key=lambda x: x["Question Number"])
            return pd.DataFrame(data)

        # Process the Answer Key File
        if answer_key_file.name.lower().endswith(".docx"):
            st.info("Extracting answer key from Word Document...")
            try:
                df_answer_key = extract_answer_key_from_docx(answer_key_file)
                
                if df_answer_key is not None:
                    st.success("Answer key extracted successfully!")
                    
                    st.write("### Answer Key Preview")
                    st.dataframe(df_answer_key.head(10))
                    
                    ak_buffer = io.BytesIO()
                    df_answer_key.to_excel(ak_buffer, index=False)
                    ak_buffer.seek(0)
                    
                    st.download_button(
                        label="⬇️ Download Converted Answer Key (Excel)",
                        data=ak_buffer,
                        file_name="Converted_Answer_Key.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.error("Could not find a valid answer key table in the Word document.")
            except Exception as e:
                st.error(f"Error processing Word document: {e}. Note: .doc files are not supported, please convert to .docx")
        
        elif answer_key_file.name.lower().endswith(".xlsx") or answer_key_file.name.lower().endswith(".xls"):
            try:
                df_answer_key = pd.read_excel(answer_key_file)
                st.success("Excel answer key loaded successfully!")
                st.write("### Answer Key Preview")
                st.dataframe(df_answer_key.head(10))
            except Exception as e:
                st.error(f"Error reading Excel file: {e}")
                
        # Process the DAT File against the Answer Key
        if df_answer_key is not None:
            if st.button("Grade DAT File"):
                if "jee" in selected_batch.lower():
                    if True:
                        st.info("Grading student DAT file against the Answer Key...")
                    
                    try:
                        # Convert df_ak answers to a dictionary: {q_num: answer}
                        # map A=1, B=2, C=3, D=4
                        ans_map = {'A':'1', 'B':'2', 'C':'3', 'D':'4'}
                        ak_dict = {}
                        for _, row in df_answer_key.iterrows():
                            q_num = int(row["Question Number"])
                            ans_raw = str(row["Answer"]).strip().upper()
                            ans_val = ans_map.get(ans_raw, ans_raw)
                            ak_dict[q_num] = ans_val
                            
                        results = []
                        
                        # Read DAT file
                        content = dat_file.read().decode('utf-8').splitlines()
                        for line in content:
                            line = line.strip()
                            if not line:
                                continue
                            
                            # Split by comma or space
                            parts = [p.strip() for p in line.replace(',', ' ').split() if p.strip()]
                            if len(parts) < 77:
                                continue # Malformed row
                                
                            student_id = parts[1]
                            responses = parts[2:77] # 75 responses
                            
                            # Sub-sections
                            phys_obj = responses[0:20]
                            phys_int = responses[20:25]
                            chem_obj = responses[25:45]
                            chem_int = responses[45:50]
                            math_obj = responses[50:70]
                            math_int = responses[70:75]
                            
                            def score_section(obj_responses, int_responses, start_q):
                                marks = 0
                                wrong = 0
                                unattempted = 0
                                
                                # Objective
                                for i, r in enumerate(obj_responses):
                                    q = start_q + i
                                    correct_ans = ak_dict.get(q, None)
                                    if r == '0':
                                        unattempted += 1
                                    elif correct_ans and r == str(correct_ans):
                                        marks += 4
                                    else:
                                        marks -= 1
                                        wrong += 1
                                        
                                # Integer
                                for i, r in enumerate(int_responses):
                                    q = start_q + 20 + i
                                    correct_ans = ak_dict.get(q, None)
                                    
                                    if "-10000" in r:
                                        unattempted += 1
                                        continue
                                        
                                    if correct_ans:
                                        try:
                                            r_float = float(r)
                                            c_float = float(correct_ans)
                                            if r_float == c_float:
                                                marks += 4
                                            else:
                                                marks -= 1
                                                wrong += 1
                                        except ValueError:
                                            if r == str(correct_ans):
                                                marks += 4
                                            else:
                                                marks -= 1
                                                wrong += 1
                                    else:
                                        unattempted += 1
                    
                                return marks, wrong, unattempted
                    
                            p_m, p_w, p_u = score_section(phys_obj, phys_int, 1)
                            c_m, c_w, c_u = score_section(chem_obj, chem_int, 26)
                            m_m, m_w, m_u = score_section(math_obj, math_int, 51)
                            
                            total = p_m + c_m + m_m
                            
                            results.append({
                                "Student Code": student_id,
                                "Physics Marks": p_m,
                                "Physics Wrong": p_w,
                                "Physics Unattempted": p_u,
                                "Chemistry Marks": c_m,
                                "Chemistry Wrong": c_w,
                                "Chemistry Unattempted": c_u,
                                "Math Marks": m_m,
                                "Math Wrong": m_w,
                                "Math Unattempted": m_u,
                                "Total Marks": total
                            })
                            
                        df_results = pd.DataFrame(results)
                        
                        # Sort by Total Marks descending
                        if not df_results.empty:
                            df_results = df_results.sort_values(by="Total Marks", ascending=False).reset_index(drop=True)
                        
                        st.success("DAT file graded successfully!")
                        st.write("### Graded Results Preview")
                        st.dataframe(df_results.head(10))
                        
                        out_buffer = io.BytesIO()
                        df_results.to_excel(out_buffer, index=False)
                        out_buffer.seek(0)
                        
                        st.download_button(
                            label="⬇️ Download Graded Results (Excel)",
                            data=out_buffer,
                            file_name="Graded_JEE_Student_Results.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    except Exception as e:
                        st.error(f"Error grading JEE DAT file: {e}")

                elif "neet" in selected_batch.lower():
                    st.info("Grading NEET student DAT file against the Answer Key...")
                    try:
                        ans_map = {'A':'1', 'B':'2', 'C':'3', 'D':'4'}
                        ak_dict = {}
                        for _, row in df_answer_key.iterrows():
                            q_num = int(row["Question Number"])
                            ans_raw = str(row["Answer"]).strip().upper()
                            ans_val = ans_map.get(ans_raw, ans_raw)
                            ak_dict[q_num] = ans_val
                            
                        results = []
                        content = dat_file.read().decode('utf-8').splitlines()
                        
                        clean_lines = [line.strip() for line in content if line.strip()]
                        
                        # 1=1, 2=2, 4=3, 8=4 mapping for NEET dat files
                        student_code_map = {'1': '1', '2': '2', '4': '3', '8': '4'}
                        
                        for i in range(0, len(clean_lines), 4):
                            if i + 3 >= len(clean_lines):
                                break
                            
                            line1 = clean_lines[i]
                            if "No.=" in line1:
                                student_id = line1.split("No.=")[-1].strip()
                            else:
                                student_id = line1.replace("No.", "").replace("=", "").strip()
                                
                            line4 = clean_lines[i+3]
                            responses_raw = line4.split()
                            
                            responses = []
                            for r in responses_raw:
                                responses.append(student_code_map.get(r, '0'))
                                
                            responses = responses[:200] + ['0'] * max(0, 200 - len(responses))
                            
                            phys_resp = responses[0:50]
                            chem_resp = responses[50:100]
                            bot_resp = responses[100:150]
                            zoo_resp = responses[150:200]
                            
                            def score_neet_section(resp_list, start_q):
                                marks = 0
                                wrong = 0
                                unattempted = 0
                                
                                for idx, r in enumerate(resp_list):
                                    q = start_q + idx
                                    correct_ans = ak_dict.get(q, None)
                                    
                                    if r == '0':
                                        unattempted += 1
                                    elif correct_ans and r == str(correct_ans):
                                        marks += 4
                                    else:
                                        marks -= 1
                                        wrong += 1
                                        
                                return marks, wrong, unattempted

                            p_m, p_w, p_u = score_neet_section(phys_resp, 1)
                            c_m, c_w, c_u = score_neet_section(chem_resp, 51)
                            b_m, b_w, b_u = score_neet_section(bot_resp, 101)
                            z_m, z_w, z_u = score_neet_section(zoo_resp, 151)
                            
                            total = p_m + c_m + b_m + z_m
                            
                            results.append({
                                "Student Code": student_id,
                                "Physics Marks": p_m,
                                "Physics Wrong": p_w,
                                "Physics Unattempted": p_u,
                                "Chemistry Marks": c_m,
                                "Chemistry Wrong": c_w,
                                "Chemistry Unattempted": c_u,
                                "Botany Marks": b_m,
                                "Botany Wrong": b_w,
                                "Botany Unattempted": b_u,
                                "Zoology Marks": z_m,
                                "Zoology Wrong": z_w,
                                "Zoology Unattempted": z_u,
                                "Total Marks": total
                            })
                            
                        df_results = pd.DataFrame(results)
                        if not df_results.empty:
                            df_results = df_results.sort_values(by="Total Marks", ascending=False).reset_index(drop=True)
                        
                        st.success("NEET DAT file graded successfully!")
                        st.write("### Graded Results Preview")
                        st.dataframe(df_results.head(10))
                        
                        out_buffer = io.BytesIO()
                        df_results.to_excel(out_buffer, index=False)
                        out_buffer.seek(0)
                        
                        st.download_button(
                            label="⬇️ Download Graded Results (Excel)",
                            data=out_buffer,
                            file_name="Graded_NEET_Results.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                    except Exception as e:
                        st.error(f"Error grading NEET DAT file: {e}")
